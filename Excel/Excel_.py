import re
import pytest
import openpyxl
from playwright.sync_api import Page, expect, sync_playwright
from funciones_excel import Funciones_Globales

url = "https://demoqa.com/text-box"
nRegistros = 11
#Reportes con Reportes Html1
# pytest Excel_.py -s -v -n2 --template=html1/index.html --report=report_excel.html
# intalar la siguiente libreria openpyxl:  pip install openpyxl

ruta_Excel = "C:/Practicas/Playwright/Playwright/Excel/datos_prueba.xlsx"
registros = 15

archivo = openpyxl.load_workbook(ruta_Excel)
def NumeroFilas(hoja):
    ac = archivo[hoja]
    return ac.max_row

def DatosColumna(hoja,fila,col):
    ac = archivo[hoja]
    col = ac.cell(int(fila),int(col))
    return col.value

# print("Numero de filas: ", NumeroFilas("Hoja1"))
# print("Datos de la columna 1, fila 2: ", DatosColumna("Hoja1",12,1))

def test_excel(set_up_excel)-> None:
    page: Page = set_up_excel
    F = Funciones_Globales(page)
    expect(page).to_have_title("DEMOQA")
    F.Scroll(0, 300)

    # Leer datos de Excel
    for fila in range(2, nRegistros):
        nombre = DatosColumna("Hoja1", fila, 1)
        email = DatosColumna("Hoja1", fila, 2)
        direccion = DatosColumna("Hoja1", fila, 3)
        estado = DatosColumna("Hoja1", fila, 4)

        F.Scroll(0, 200)       

        # Interactuar con la página
        F.Texto("#userName", nombre)
        F.Texto("#userEmail", email)
        F.Texto("#currentAddress", direccion)
        F.Texto("#permanentAddress", estado)

        F.TomarScreenshot(f"evidencia_fila_{fila}.png")

        F.Click("#submit")
        F.Esperar(.6)
        page.goto(url)
        page.set_default_timeout(2000) 
        F.Esperar(.6)
        

        
        
